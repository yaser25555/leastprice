#!/usr/bin/env python3
"""
Import coupon rows from an Excel workbook into Firestore.

Expected source columns can be Arabic or English variants of:
  - coupon code
  - discount / discount percent
  - store / advertiser
  - expiry / validity date

Required environment variables for real imports:
  GOOGLE_APPLICATION_CREDENTIALS=/abs/path/service-account.json

Optional:
  FIREBASE_PROJECT_ID=leastprice-yaser
  COUPONS_COLLECTION=coupons

Example:
  python scripts/import_coupons.py --file cc-coupons.xlsx --dry-run
  python scripts/import_coupons.py --file cc-coupons.xlsx
"""

from __future__ import annotations

import argparse
import os
import re
import sys
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover - runtime guard
    raise SystemExit(
        "Missing dependency 'openpyxl'. Install it with: python -m pip install openpyxl"
    ) from exc

try:
    import firebase_admin
    from firebase_admin import credentials, firestore
except ImportError as exc:  # pragma: no cover - runtime guard
    raise SystemExit(
        "Missing dependency 'firebase-admin'. Install it with: python -m pip install firebase-admin"
    ) from exc


DEFAULT_PROJECT_ID = os.getenv("FIREBASE_PROJECT_ID", "leastprice-yaser").strip() or "leastprice-yaser"
DEFAULT_COLLECTION = os.getenv("COUPONS_COLLECTION", "coupons").strip() or "coupons"

HEADER_ALIASES = {
    "code": {
        "coupon code",
        "code",
        "coupon",
        "promo code",
        "discount code",
        "كود",
        "كود الخصم",
        "رمز الخصم",
        "الكود",
    },
    "discount": {
        "discount",
        "discount percent",
        "discount percentage",
        "discount %",
        "offer",
        "نسبه الخصم",
        "نسبة الخصم",
        "خصم",
        "الخصم",
    },
    "store": {
        "store",
        "advertiser",
        "merchant",
        "brand",
        "platform",
        "المتجر",
        "المنصة",
        "المعلن",
    },
    "expiry": {
        "expiry",
        "expiry date",
        "expires at",
        "valid until",
        "validity",
        "end date",
        "تاريخ الانتهاء",
        "الانتهاء",
        "الصلاحية",
        "صالح حتى",
    },
    "title": {
        "title",
        "name",
        "offer title",
        "عنوان",
        "الاسم",
    },
    "description": {
        "description",
        "notes",
        "details",
        "الوصف",
        "ملاحظات",
    },
}


@dataclass(frozen=True)
class ParsedCoupon:
    document_id: str
    payload: Dict[str, Any]


def normalize_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).lower()


def normalize_store_id(value: Any) -> str:
    normalized = normalize_text(value)
    if "noon" in normalized or "نون" in normalized:
        return "noon"
    if "namshi" in normalized or "نمشي" in normalized:
        return "namshi"
    return re.sub(r"[^a-z0-9]+", "", normalized)


def store_name_for_id(store_id: str) -> str:
    if store_id == "noon":
      return "Noon"
    if store_id == "namshi":
      return "Namshi"
    return store_id or "Online store"


def parse_discount_label(value: Any) -> str:
    text = str(value or "").strip()
    if text:
        return text
    return "Special discount"


def parse_discount_percent(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    match = re.search(r"([0-9]+(?:\.[0-9]+)?)", str(value).replace(",", "."))
    if not match:
        return None
    return float(match.group(1))


def parse_datetime(value: Any) -> Optional[datetime]:
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, datetime):
        return value if value.tzinfo else value.replace(tzinfo=UTC)

    text = str(value).strip()
    candidates = [
        "%Y-%m-%d",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%dT%H:%M:%SZ",
        "%d/%m/%Y",
        "%m/%d/%Y",
        "%d-%m-%Y",
    ]
    for fmt in candidates:
        try:
            parsed = datetime.strptime(text, fmt)
            return parsed.replace(tzinfo=UTC)
        except ValueError:
            continue
    return None


def canonical_header(header_value: Any) -> Optional[str]:
    normalized = normalize_text(header_value)
    if not normalized:
        return None
    for canonical, aliases in HEADER_ALIASES.items():
        if normalized in aliases:
            return canonical
    return None


def build_header_map(header_row: Iterable[Any]) -> Dict[int, str]:
    header_map: Dict[int, str] = {}
    for index, header_value in enumerate(header_row):
        canonical = canonical_header(header_value)
        if canonical:
            header_map[index] = canonical
    return header_map


def parse_row(row: Iterable[Any], header_map: Dict[int, str], row_number: int) -> ParsedCoupon:
    values: Dict[str, Any] = {}
    for index, cell_value in enumerate(row):
        canonical = header_map.get(index)
        if canonical:
            values[canonical] = cell_value

    code = str(values.get("code") or "").strip()
    if not code:
        raise ValueError("missing coupon code")

    store_id = normalize_store_id(values.get("store"))
    if store_id not in {"noon", "namshi"}:
        raise ValueError(f"unsupported store: {values.get('store')!r}")

    expires_at = parse_datetime(values.get("expiry"))
    if expires_at is None:
        raise ValueError("missing or invalid expiry date")

    discount_label = parse_discount_label(values.get("discount"))
    discount_percent = parse_discount_percent(values.get("discount"))

    document_id = f"{store_id}_{re.sub(r'[^A-Za-z0-9]+', '_', code).strip('_').lower()}"
    payload = {
        "code": code,
        "storeId": store_id,
        "storeName": store_name_for_id(store_id),
        "discountLabel": discount_label,
        "expiresAt": expires_at,
        "active": True,
        "title": str(values.get("title") or "").strip() or None,
        "description": str(values.get("description") or "").strip() or None,
        "discountPercent": discount_percent,
        "sourceRowNumber": row_number,
        "updatedAt": firestore.SERVER_TIMESTAMP,
    }
    return ParsedCoupon(document_id=document_id, payload=payload)


def initialize_firestore(project_id: str):
    if not firebase_admin._apps:  # type: ignore[attr-defined]
        credentials_path = os.getenv("GOOGLE_APPLICATION_CREDENTIALS", "").strip()
        if credentials_path:
            firebase_admin.initialize_app(credentials.Certificate(credentials_path), {"projectId": project_id})
        else:
            firebase_admin.initialize_app(options={"projectId": project_id})
    return firestore.client()


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import coupons from Excel into Firestore.")
    parser.add_argument("--file", default="cc-coupons.xlsx", help="Path to the source workbook.")
    parser.add_argument("--sheet", default="", help="Optional worksheet name.")
    parser.add_argument("--project-id", default=DEFAULT_PROJECT_ID, help="Firebase project id.")
    parser.add_argument("--collection", default=DEFAULT_COLLECTION, help="Firestore collection name.")
    parser.add_argument("--dry-run", action="store_true", help="Parse rows without writing to Firestore.")
    return parser.parse_args()


def main() -> int:
    args = parse_arguments()
    workbook_path = Path(args.file).expanduser().resolve()
    if not workbook_path.exists():
        print(f"Workbook not found: {workbook_path}", file=sys.stderr)
        return 1

    workbook = openpyxl.load_workbook(workbook_path, data_only=True)
    worksheet = workbook[args.sheet] if args.sheet else workbook[workbook.sheetnames[0]]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        print("Workbook is empty.", file=sys.stderr)
        return 1

    header_map = build_header_map(rows[0])
    required_headers = {"code", "discount", "store", "expiry"}
    missing_headers = sorted(required_headers - set(header_map.values()))
    if missing_headers:
        print(
            f"Missing required columns: {', '.join(missing_headers)}. "
            "The current workbook does not match the expected coupon schema.",
            file=sys.stderr,
        )
        return 1

    parsed: List[ParsedCoupon] = []
    skipped: List[str] = []

    for row_number, row in enumerate(rows[1:], start=2):
        if not any(cell is not None and str(cell).strip() for cell in row):
            continue
        try:
            parsed.append(parse_row(row, header_map, row_number))
        except ValueError as exc:
            skipped.append(f"Row {row_number}: {exc}")

    print(f"Workbook: {workbook_path}")
    print(f"Worksheet: {worksheet.title}")
    print(f"Valid coupons: {len(parsed)}")
    print(f"Skipped rows: {len(skipped)}")
    if skipped:
        print("Skip details:")
        for line in skipped[:20]:
            print(f"  - {line}")

    if not parsed:
        print("No valid coupon rows were found, so nothing was imported.", file=sys.stderr)
        return 1

    if args.dry_run:
        print("Dry run complete. No Firestore writes were made.")
        return 0

    database = initialize_firestore(args.project_id)
    collection = database.collection(args.collection)

    imported = 0
    for coupon in parsed:
        payload = dict(coupon.payload)
        payload["createdAt"] = firestore.SERVER_TIMESTAMP
        collection.document(coupon.document_id).set(payload, merge=True)
        imported += 1

    print(f"Imported {imported} coupons into '{args.collection}'.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
